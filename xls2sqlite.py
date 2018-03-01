import sqlite3
from openpyxl import load_workbook
import re
from bs4 import BeautifulSoup

from duendecat.dir import database_path
from duendecat.common.furigana import kanji


def main():
    name = 'JLPT'
    xlsx_name = name + '.xlsx'
    db_name = name + '.db'
    with sqlite3.connect(db_name) as conn:
        wb = load_workbook(database_path(xlsx_name))
        for sheet in wb.sheetnames:
            conn.execute('''
                CREATE TABLE {} (
                    ID INTEGER PRIMARY KEY AUTOINCREMENT,
                    lang TEXT,
                    reading TEXT,
                    en TEXT,
                    grammar_point TEXT,
                    grammar_level TEXT,
                    char_level TEXT
                );
            '''.format(slugify(sheet)))

            for i, row in enumerate(wb[sheet]):
                if i == 0:
                    continue
                lang, reading, en, grammar_point, grammar_level, char_level = \
                    getcol(wb, xlsx_name, sheet, row)
                print('\n'.join(getcol(wb, xlsx_name, sheet, row)))
                conn.execute('''
                    INSERT INTO {} (lang, reading, en, grammar_point, grammar_level, char_level)
                    VALUES (?, ?, ?, ?, ?, ?)
                '''.format(sheet), (lang, reading, en, grammar_point, grammar_level, char_level))


def getcol(wb, xlsx_name, sheet, row):
    def num(c):
        return ord(c) - ord('A')

    def noHtml(sentence):
        return BeautifulSoup(sentence, 'html.parser').text

    def langSen(sentence):
        return kanji(noHtml(sentence))

    def val(c):
        result = str(row[num(c)].value)
        if result is None:
            return ''
        else:
            return result

    def getHSK():
        if sheet[0] in ('A', 'B', 'C'):
            lang = val('A')
            reading = val('B')
            en = val('C')
            grammar_point = val('F') + '\n' + val('G')
            grammar_level = sheet
            char_level = val('K')
        elif sheet == 'SpoonFed':
            lang = val('C')
            reading = val('B')
            en = val('A')
            grammar_point = 'SpoonFed Chinese sentences'
            grammar_level = 'SpoonFed order: ' + val('E')
            char_level = val('E')
        else:
            raise ValueError('Sheet not found')
        return langSen(lang), \
               noHtml(reading), \
               noHtml(en), \
               grammar_point, grammar_level, char_level

    def getJLPT():
        if sheet[0] == 'N':
            lang = val('A')
            reading = val('F')
            en = val('B')
            grammar_point = val('G')
            grammar_level = val('D')
            char_level = val('H')
        elif sheet[0:2] == 'WK':
            lang = val('C')
            reading = val('C')
            en = val('D')
            grammar_point = 'WaniKani vocab: ' + val('B')
            grammar_level = 'Appears in context sentence level: ' + val('A')
            char_level = val('E')
        elif sheet == 'Kana':
            lang = val('E')
            reading = val('E')
            en = val('F')
            grammar_point = 'Kana vocab: ' + val('B')
            grammar_level = 'Vocab frequency: ' + val('G')
            char_level = val('J')
        elif sheet[0:3] == 'Set':
            lang = val('F')
            reading = val('F')
            en = val('G')
            grammar_point = sheet + ' vocab: ' + val('B')
            grammar_level = 'Vocab frequency: ' + val('H')
            char_level = val('K')
        else:
            raise ValueError('Sheet not found')
        return langSen(lang), \
               noHtml(reading), \
               noHtml(en), \
               grammar_point, grammar_level, char_level

    return getJLPT() if xlsx_name == 'JLPT.xlsx' else getHSK()


def slugify(text, lower=1):
    if lower == 1:
        text = text.strip().lower()
    text = re.sub(r'[^\w _-]+', '', text)
    text = re.sub(r'[- ]+', '_', text)
    return text


if __name__ == '__main__':
    main()