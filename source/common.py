import sys, os
from google_speech import Speech
from threading import Thread

from source import furigana
from openpyxl import *
import os, sys
from time import sleep
from bs4 import BeautifulSoup
import logging
from random import choice

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

class Data():
	def __init__(self, **param):
		logging.debug(param)

		self.reverse = param['is_reverse']
		self.lang = param['lang']
		self.is_speak = param['speak']
		self.is_speaking = False
		self.speech_engine = param['speech_engine']

		if self.lang == 'cn':
			filename = 'HSK.xlsx'
			self.character_type = 'Hanzi'
		else:
			filename = 'JLPT.xlsx'
			self.character_type = 'Kanji'

		application_path = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
		if getattr(sys, 'frozen', False):
			application_path = os.path.dirname(sys.executable)
		os.chdir(application_path)

		logging.debug('CWD: ' + os.getcwd())
		logging.debug('Loading Excel')
		self.wb = load_workbook(os.path.join('database',filename))
		logging.debug('Excel loaded\n')

		if param['sheet'] == 'any':
			self.sheet = choice(self.wb.get_sheet_names())
		else:
			self.sheet = param['sheet']
		self.worksheet = self.wb[self.sheet]
		self.max_row = self.worksheet.max_row

		self.col = dict()
		if self.sheet[0] == 'N':
			self.col['lang'] = 'A'
			self.col['reading'] = 'F'
			self.col['en'] = 'B'
			self.col['grammar_point'] = 'G'
			self.col['grammar_level'] = 'D'
			self.col['kanji_level'] = 'H'
		elif self.sheet[0:2] == 'WK':
			self.col['lang'] = 'C'
			self.col['reading'] = 'C'
			self.col['en'] = 'D'
			self.col['grammar_point'] = 'WaniKani vocab: B'
			self.col['grammar_level'] = 'Appears in context sentence level: A'
			self.col['kanji_level'] = 'E'
		elif self.sheet == 'Kana':
			self.col['lang'] = 'E'
			self.col['reading'] = 'E'
			self.col['en'] = 'F'
			self.col['grammar_point'] = self.sheet + ' vocab: B'
			self.col['grammar_level'] = 'Vocab frequency: G'
			self.col['kanji_level'] = 'J'
		elif self.sheet[0:3] == 'Set':
			self.col['lang'] = 'F'
			self.col['reading'] = 'F'
			self.col['en'] = 'G'
			self.col['grammar_point'] = self.sheet + ' vocab: B'
			self.col['grammar_level'] = 'Vocab frequency: H'
			self.col['kanji_level'] = 'K'
		elif self.sheet[0] in ('A', 'B', 'C'):
			self.col['lang'] = 'A'
			self.col['reading'] = 'B'
			self.col['en'] = 'C'
			self.col['grammar_point'] = 'FG'
			self.col['grammar_level'] = self.sheet
			self.col['kanji_level'] = 'K'
		elif self.sheet == 'SpoonFed':
			self.col['lang'] = 'C'
			self.col['reading'] = 'B'
			self.col['en'] = 'A'
			self.col['grammar_point'] = 'SpoonFed Chinese sentences'
			self.col['grammar_level'] = 'SpoonFed order: E'
			self.col['kanji_level'] = 'E'

	def getLangSen(self, row):
		raw = self.worksheet[self.col['lang'] + str(row)].value
		return furigana.kanji(BeautifulSoup(raw, 'lxml').text)

	def getReadingSen(self, row):
		raw = self.worksheet[self.col['reading'] + str(row)].value
		return BeautifulSoup(raw, 'lxml').text

	def getEnSen(self, row):
		raw = self.worksheet[self.col['en'] + str(row)].value
		return BeautifulSoup(raw, 'lxml').text

	def getGrammar(self, row):
		grammar_point = self.formatGrammar(self.col['grammar_point'], row)
		grammar_level = self.formatGrammar(self.col['grammar_level'], row)

		raw = grammar_point + '\n' + grammar_level
		return furigana.kanji(BeautifulSoup(raw, 'lxml').text)

	def getLevel(self, row):
		level = self.formatGrammar(self.col['kanji_level'], row)
		return self.character_type + ' level ' + str(level)

	def getLatterPortion(self, row):
		output = []
		output += [self.getReadingSen(row)]
		output += [self.getEnSen(row)]
		output += [self.getGrammar(row)]
		if self.sheet != 'SpoonFed':
			output += [self.getLevel(row)]

		return '\n'.join(output)

	def getLatterPortionReversed(self, row):
		output = []
		output += [self.getLangSen(row)]
		output += [self.getReadingSen(row)]
		output += [self.getGrammar(row)]
		if self.sheet != 'SpoonFed':
			output += [self.getLevel(row)]

		return '\n'.join(output)

	def formatGrammar(self, grammar, row):
		logging.debug(grammar)
		logging.debug(row)
		if len(grammar) == 1:
			grammar = self.worksheet[grammar + str(row)].value
		else:
			if ':' not in grammar:
				pass
			else:
				grammar = grammar[:-1] + str(self.worksheet[grammar[-1] + str(row)].value)

		return grammar

	def getData(self, row, reverse=False):
		if not reverse:
			first, last = self.getLangSen(row), self.getLatterPortion(row)
		else:
			first, last = self.getEnSen(row), self.getLatterPortionReversed(row)

		return first, last

	def getMaxLevelRow(self, max_level):
		if self.sheet == 'SpoonFed':
			max_level = max_level * self.worksheet.max_row / 60
		for row in range(2, self.worksheet.max_row):
			level = self.worksheet[self.col['kanji_level'] + str(row)].value
			if level > max_level:
				return row-1

		return row+1

	def speak(self, sentence_type, row, sleep=False):
		if not self.is_speak:
			return
		if self.reverse:
			if sentence_type == 'top':
				sentence_type = 'bottom'
			else:
				sentence_type = 'top'
		if sentence_type == 'top':
			self.sayLang(self.getLangSen(row), sleep)
		else:
			self.sayEn(self.getEnSen(row), sleep)

	def sayLang(self, sentence, sleep):
		self.say(sentence, self.lang, sleep)

	def sayEn(self, sentence, sleep):
		self.say(sentence, 'en', sleep)

	def say(self, sentence, lang, sleep):
		if lang == 'en':
			speaker = 'alex'
		elif lang == 'jp':
			lang = 'ja'
			speaker = 'kyoko'
		elif lang == 'cn':
			lang = 'zh-CN'
			speaker = 'ting-ting'

		def myfunc():
			self.is_speaking = True
			
			if self.speech_engine == 'not_set':
				if sys.platform == 'darwin':
					self.speech_engine = 'mac'
				else:
					self.speech_engine == 'google'

			if self.speech_engine == 'mac':
				os.system('say -v {} "{}"'.format(speaker, sentence))
			elif self.speech_engine == 'google':
				speech = Speech(sentence, lang)
				speech.play(tuple())

			self.is_speaking = False

		if not sleep:
			t = Thread(target=myfunc)
			t.start()
		else:
			myfunc()

def printAnything(anything):
	sys.stdout.buffer.write((repr(anything) + '\n').encode('utf8'))

def printText(text):
	sys.stdout.buffer.write((str(text) + '\n').encode('utf8'))